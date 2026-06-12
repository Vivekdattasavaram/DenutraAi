"""
Convert DAST report.json into a formatted Excel (.xlsx) report.
"""
import json, os
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TEST_DIR = os.path.dirname(SCRIPT_DIR)
REPORT_JSON = os.path.join(TEST_DIR, "reports", "report.json")
OUTPUT_XLSX = os.path.join(TEST_DIR, "reports", f"DAST_Security_Report_{datetime.now().strftime('%Y-%m-%dT%H-%M-%S')}.xlsx")

with open(REPORT_JSON, "r", encoding="utf-8") as f:
    results = json.load(f)

wb = Workbook()

# ─── Colors ───────────────────────────────────────────────────────
DARK_BG      = PatternFill("solid", fgColor="1B1F3B")
HEADER_BG    = PatternFill("solid", fgColor="2D3250")
PASS_BG      = PatternFill("solid", fgColor="1A3C34")
CRITICAL_BG  = PatternFill("solid", fgColor="5C1A1A")
HIGH_BG      = PatternFill("solid", fgColor="5C3A1A")
MEDIUM_BG    = PatternFill("solid", fgColor="5C5C1A")
LOW_BG       = PatternFill("solid", fgColor="1A3C5C")
ROW_EVEN     = PatternFill("solid", fgColor="232946")
ROW_ODD      = PatternFill("solid", fgColor="1B1F3B")

WHITE_FONT   = Font(color="EAEAEA", size=11)
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=12)
TITLE_FONT   = Font(color="FFFFFF", bold=True, size=16)
SUBTITLE_FONT= Font(color="B0BEC5", size=11, italic=True)
PASS_FONT    = Font(color="4CAF50", bold=True, size=11)
FAIL_FONT    = Font(color="FF5252", bold=True, size=11)
SEVERITY_FONTS = {
    "CRITICAL": Font(color="FF1744", bold=True, size=11),
    "HIGH":     Font(color="FF9100", bold=True, size=11),
    "MEDIUM":   Font(color="FFEA00", bold=True, size=11),
    "LOW":      Font(color="40C4FF", bold=True, size=11),
    "PASS":     Font(color="69F0AE", size=11),
}
SEVERITY_FILLS = {
    "CRITICAL": CRITICAL_BG,
    "HIGH":     HIGH_BG,
    "MEDIUM":   MEDIUM_BG,
    "LOW":      LOW_BG,
    "PASS":     PASS_BG,
}

thin_border = Border(
    left=Side(style="thin", color="3D4260"),
    right=Side(style="thin", color="3D4260"),
    top=Side(style="thin", color="3D4260"),
    bottom=Side(style="thin", color="3D4260"),
)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ═════════════════════════════════════════════════════════════════
# SHEET 1: Executive Summary
# ═════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Executive Summary"
ws1.sheet_properties.tabColor = "1B1F3B"

# Set column widths
ws1.column_dimensions["A"].width = 35
ws1.column_dimensions["B"].width = 25
ws1.column_dimensions["C"].width = 20
ws1.column_dimensions["D"].width = 20

# Title
ws1.merge_cells("A1:D1")
ws1["A1"] = "DAST Security Scan Report — Dentura AI"
ws1["A1"].font = TITLE_FONT
ws1["A1"].fill = DARK_BG
ws1["A1"].alignment = center
ws1.row_dimensions[1].height = 40

ws1.merge_cells("A2:D2")
ws1["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  Target: http://localhost:8000"
ws1["A2"].font = SUBTITLE_FONT
ws1["A2"].fill = DARK_BG
ws1["A2"].alignment = center

# Summary stats
total_tests = len(results)
findings = [r for r in results if r["finding"]]
total_findings = len(findings)
sev_counts = {}
for r in findings:
    sev_counts[r["severity"]] = sev_counts.get(r["severity"], 0) + 1

summary_data = [
    ("Metric", "Value", "Status", ""),
    ("Endpoints Discovered", "26", "", ""),
    ("Total Tests Executed", str(total_tests), "", ""),
    ("Total Findings", str(total_findings), "NEEDS ATTENTION" if total_findings > 0 else "ALL CLEAR", ""),
    ("CRITICAL Findings", str(sev_counts.get("CRITICAL", 0)), "PASS" if sev_counts.get("CRITICAL", 0) == 0 else "FAIL", ""),
    ("HIGH Findings", str(sev_counts.get("HIGH", 0)), "PASS" if sev_counts.get("HIGH", 0) == 0 else "FAIL", ""),
    ("MEDIUM Findings", str(sev_counts.get("MEDIUM", 0)), "PASS" if sev_counts.get("MEDIUM", 0) == 0 else "FAIL", ""),
    ("LOW Findings", str(sev_counts.get("LOW", 0)), "PASS" if sev_counts.get("LOW", 0) == 0 else "FAIL", ""),
]

for i, row_data in enumerate(summary_data):
    row = i + 4
    for j, val in enumerate(row_data):
        cell = ws1.cell(row=row, column=j + 1, value=val)
        cell.border = thin_border
        if i == 0:
            cell.font = HEADER_FONT
            cell.fill = HEADER_BG
            cell.alignment = center
        else:
            cell.fill = ROW_EVEN if i % 2 == 0 else ROW_ODD
            cell.font = WHITE_FONT
            cell.alignment = center
            if j == 2:
                if val == "PASS":
                    cell.font = PASS_FONT
                elif val == "FAIL":
                    cell.font = FAIL_FONT
                elif val == "NEEDS ATTENTION":
                    cell.font = Font(color="FF9100", bold=True, size=11)

# Category breakdown
cat_row = 14
ws1.merge_cells(f"A{cat_row}:D{cat_row}")
ws1[f"A{cat_row}"] = "Test Category Breakdown"
ws1[f"A{cat_row}"].font = Font(color="FFFFFF", bold=True, size=14)
ws1[f"A{cat_row}"].fill = HEADER_BG
ws1[f"A{cat_row}"].alignment = center
ws1.row_dimensions[cat_row].height = 30

cat_headers = ["Category", "Tests Run", "Findings", "Result"]
cat_row += 1
for j, h in enumerate(cat_headers):
    cell = ws1.cell(row=cat_row, column=j + 1, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_BG
    cell.alignment = center
    cell.border = thin_border

categories = {}
for r in results:
    cat = r["test_category"]
    if cat not in categories:
        categories[cat] = {"total": 0, "findings": 0}
    categories[cat]["total"] += 1
    if r["finding"]:
        categories[cat]["findings"] += 1

for i, (cat, data) in enumerate(categories.items()):
    row = cat_row + 1 + i
    result = "PASS" if data["findings"] == 0 else "FAIL"
    vals = [cat, data["total"], data["findings"], result]
    for j, val in enumerate(vals):
        cell = ws1.cell(row=row, column=j + 1, value=val)
        cell.border = thin_border
        cell.fill = ROW_EVEN if i % 2 == 0 else ROW_ODD
        cell.font = WHITE_FONT
        cell.alignment = center
        if j == 3:
            cell.font = PASS_FONT if val == "PASS" else FAIL_FONT

# ═════════════════════════════════════════════════════════════════
# SHEET 2: All Test Results
# ═════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("All Test Results")
ws2.sheet_properties.tabColor = "2D3250"

columns = [
    ("S.No", 6),
    ("Endpoint", 42),
    ("Method", 10),
    ("Role", 18),
    ("HTTP Status", 14),
    ("Expected", 14),
    ("Finding", 10),
    ("Severity", 12),
    ("Response (ms)", 14),
    ("Test Category", 20),
    ("Note", 60),
    ("Timestamp", 22),
]

for j, (col_name, width) in enumerate(columns):
    col_letter = get_column_letter(j + 1)
    ws2.column_dimensions[col_letter].width = width
    cell = ws2.cell(row=1, column=j + 1, value=col_name)
    cell.font = HEADER_FONT
    cell.fill = HEADER_BG
    cell.alignment = center
    cell.border = thin_border

ws2.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
ws2.freeze_panes = "A2"

for i, r in enumerate(results):
    row = i + 2
    vals = [
        i + 1,
        r["endpoint"],
        r["method"],
        r["role"],
        r["status"],
        r["expected_status"],
        "YES" if r["finding"] else "NO",
        r["severity"],
        r["response_time_ms"],
        r["test_category"],
        r["note"],
        r["timestamp"][:19].replace("T", " "),
    ]
    for j, val in enumerate(vals):
        cell = ws2.cell(row=row, column=j + 1, value=val)
        cell.border = thin_border
        cell.alignment = left_wrap if j in (1, 10) else center
        
        bg = ROW_EVEN if i % 2 == 0 else ROW_ODD
        cell.fill = bg
        cell.font = WHITE_FONT
        
        # Color the finding column
        if j == 6:
            if val == "YES":
                cell.font = FAIL_FONT
                cell.fill = CRITICAL_BG
            else:
                cell.font = PASS_FONT
        
        # Color the severity column
        if j == 7:
            cell.font = SEVERITY_FONTS.get(val, WHITE_FONT)
            cell.fill = SEVERITY_FILLS.get(val, bg)

# ═════════════════════════════════════════════════════════════════
# SHEET 3: Findings Only
# ═════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Findings Only")
ws3.sheet_properties.tabColor = "FF5252"

finding_cols = [
    ("S.No", 6),
    ("Severity", 12),
    ("Category", 20),
    ("Endpoint", 42),
    ("Method", 10),
    ("Role", 18),
    ("HTTP Status", 14),
    ("Note / Description", 70),
    ("Recommendation", 50),
]

for j, (col_name, width) in enumerate(finding_cols):
    col_letter = get_column_letter(j + 1)
    ws3.column_dimensions[col_letter].width = width
    cell = ws3.cell(row=1, column=j + 1, value=col_name)
    cell.font = HEADER_FONT
    cell.fill = PatternFill("solid", fgColor="4A0000")
    cell.alignment = center
    cell.border = thin_border

RECOMMENDATIONS = {
    "Hardcoded_Creds": "Move secrets to environment variables. Add .env to .gitignore. Rotate exposed credentials immediately.",
    "Rate_Limiting": "Install slowapi (pip install slowapi) and add rate limiting middleware to auth endpoints (5 attempts/min).",
    "CORS_Config": "Replace allow_origins=['*'] with your specific frontend domain(s) before production deployment.",
    "AuthN_Bypass": "Ensure all non-public endpoints use Depends(get_current_user).",
    "AuthZ_PrivEsc": "Verify admin role check using require_admin dependency on all /api/admin/* routes.",
    "Token_Tampering": "Ensure JWT signature verification is enforced. Never accept unsigned or re-encoded tokens.",
    "Injection_SQLi": "Use parameterized queries (SQLAlchemy ORM handles this). Add input validation with Pydantic.",
    "IDOR": "Always filter queries by current_user.id. Never trust client-provided user IDs.",
}

findings_only = [r for r in results if r["finding"]]
for i, r in enumerate(findings_only):
    row = i + 2
    rec = RECOMMENDATIONS.get(r["test_category"], "Review and remediate.")
    vals = [
        i + 1,
        r["severity"],
        r["test_category"],
        r["endpoint"],
        r["method"],
        r["role"],
        r["status"],
        r["note"],
        rec,
    ]
    for j, val in enumerate(vals):
        cell = ws3.cell(row=row, column=j + 1, value=val)
        cell.border = thin_border
        cell.alignment = left_wrap if j in (3, 7, 8) else center
        
        bg = ROW_EVEN if i % 2 == 0 else ROW_ODD
        cell.fill = bg
        cell.font = WHITE_FONT
        
        if j == 1:
            cell.font = SEVERITY_FONTS.get(val, WHITE_FONT)
            cell.fill = SEVERITY_FILLS.get(val, bg)

# ═════════════════════════════════════════════════════════════════
# SHEET 4: Endpoint Catalog
# ═════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Endpoint Catalog")
ws4.sheet_properties.tabColor = "4CAF50"

ep_cols = [("S.No", 6), ("Method", 10), ("Path", 50), ("Access Level", 18), ("Module", 18)]
for j, (col_name, width) in enumerate(ep_cols):
    col_letter = get_column_letter(j + 1)
    ws4.column_dimensions[col_letter].width = width
    cell = ws4.cell(row=1, column=j + 1, value=col_name)
    cell.font = HEADER_FONT
    cell.fill = PatternFill("solid", fgColor="1B5E20")
    cell.alignment = center
    cell.border = thin_border

ENDPOINTS = [
    ("GET", "/", "Public", "Root"),
    ("POST", "/api/auth/register", "Public", "Auth"),
    ("POST", "/api/auth/verify-otp", "Public", "Auth"),
    ("POST", "/api/auth/login", "Public", "Auth"),
    ("POST", "/api/auth/forgot-password", "Public", "Auth"),
    ("POST", "/api/auth/reset-password", "Public", "Auth"),
    ("POST", "/api/auth/change-password", "User", "Auth"),
    ("GET", "/api/auth/me", "User", "Auth"),
    ("POST", "/api/assessment/start", "User", "Assessment"),
    ("POST", "/api/assessment/answer", "User", "Assessment"),
    ("GET", "/api/assessment/question/next", "User", "Assessment"),
    ("POST", "/api/assessment/submit/{id}", "User", "Assessment"),
    ("GET", "/api/assessment/history", "User", "Assessment"),
    ("GET", "/api/learning/dashboard", "User", "Learning"),
    ("GET", "/api/learning/curriculum/path", "User", "Learning"),
    ("GET", "/api/learning/curriculum/module/{id}", "User", "Learning"),
    ("POST", "/api/learning/curriculum/module/{id}/complete", "User", "Learning"),
    ("POST", "/api/chatbot/message", "User", "Chatbot"),
    ("GET", "/api/notifications", "User", "Notifications"),
    ("GET", "/api/notifications/unread-count", "User", "Notifications"),
    ("POST", "/api/notifications/read/{id}", "User", "Notifications"),
    ("POST", "/api/notifications/read-all", "User", "Notifications"),
    ("GET", "/api/admin/dashboard", "Admin", "Admin"),
    ("GET", "/api/admin/users", "Admin", "Admin"),
    ("GET", "/api/admin/analytics", "Admin", "Admin"),
    ("GET", "/api/admin/notification-analytics", "Admin", "Admin"),
]

for i, (method, path, access, module) in enumerate(ENDPOINTS):
    row = i + 2
    vals = [i + 1, method, path, access, module]
    for j, val in enumerate(vals):
        cell = ws4.cell(row=row, column=j + 1, value=val)
        cell.border = thin_border
        cell.fill = ROW_EVEN if i % 2 == 0 else ROW_ODD
        cell.font = WHITE_FONT
        cell.alignment = center if j != 2 else left_wrap
        if j == 3:
            if val == "Public":
                cell.font = Font(color="69F0AE", size=11)
            elif val == "User":
                cell.font = Font(color="40C4FF", size=11)
            elif val == "Admin":
                cell.font = Font(color="FF9100", bold=True, size=11)

# Save
wb.save(OUTPUT_XLSX)
print(f"\nExcel report saved: {OUTPUT_XLSX}")
